"""Big batch of extra capabilities so Ember can match (and exceed) what a person does on a PC.
~30 new tools: email, http, pdf, excel, translation, search, calc, system, hashing, qr, network, etc."""
from __future__ import annotations

import ast
import base64
import hashlib
import io
import json
import math
import operator
import os
import secrets
import socket
import string
import subprocess
import sys
import time
import urllib.parse
import uuid
from datetime import datetime, timezone, timedelta
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from pathlib import Path

import requests


# ---------------------------------------------------------------------------
# Network / web
# ---------------------------------------------------------------------------
def http_get(url: str, headers: dict | None = None, timeout: int = 15) -> dict:
    """GET an HTTP URL and return status, headers, text body."""
    try:
        r = requests.get(url, headers=headers or {}, timeout=timeout)
        return {
            "ok": r.ok,
            "status": r.status_code,
            "url": r.url,
            "headers": dict(r.headers),
            "text": r.text[:20000],
            "json": (r.json() if r.headers.get("content-type", "").startswith("application/json") else None),
        }
    except Exception as e:
        return {"ok": False, "error": str(e)}


def http_post(url: str, data: dict | str | None = None, json_body: dict | None = None,
              headers: dict | None = None, timeout: int = 15) -> dict:
    """POST data (form-encoded if dict, raw if str) or JSON body to a URL."""
    try:
        kwargs = {"headers": headers or {}, "timeout": timeout}
        if json_body is not None:
            kwargs["json"] = json_body
        elif data is not None:
            kwargs["data"] = data
        r = requests.post(url, **kwargs)
        return {
            "ok": r.ok, "status": r.status_code, "url": r.url,
            "text": r.text[:20000],
        }
    except Exception as e:
        return {"ok": False, "error": str(e)}


def download_file(url: str, destination: str, timeout: int = 60) -> dict:
    """Stream a URL to a local file."""
    try:
        dst = Path(destination).expanduser()
        dst.parent.mkdir(parents=True, exist_ok=True)
        with requests.get(url, stream=True, timeout=timeout) as r:
            r.raise_for_status()
            total = 0
            with dst.open("wb") as f:
                for chunk in r.iter_content(chunk_size=64 * 1024):
                    if chunk:
                        f.write(chunk)
                        total += len(chunk)
        result = {"ok": True, "path": str(dst), "bytes": total,
                  "size_mb": round(total / (1024 * 1024), 2)}
        # Scan the freshly downloaded file; quarantine it if it is malicious.
        try:
            import antivirus
            gate = antivirus.gate_download(str(dst))
            if gate.get("scanned"):
                result["security"] = {k: gate[k] for k in ("verdict", "reasons", "engines")
                                      if k in gate}
                if gate.get("blocked"):
                    result["blocked"] = True
                    result["path"] = None
                    result["quarantined"] = gate.get("handled")
                    result["warning"] = ("Downloaded file was malicious and has been "
                                         "quarantined/removed.")
        except Exception:
            pass
        return result
    except Exception as e:
        return {"ok": False, "error": str(e)}


def public_ip() -> dict:
    try:
        r = requests.get("https://api.ipify.org?format=json", timeout=8)
        return {"ok": True, "ip": r.json().get("ip")}
    except Exception as e:
        return {"ok": False, "error": str(e)}


def dns_lookup(host: str) -> dict:
    try:
        info = socket.getaddrinfo(host, None)
        addrs = sorted({i[4][0] for i in info})
        return {"ok": True, "host": host, "addresses": addrs}
    except Exception as e:
        return {"ok": False, "error": str(e)}


def network_ping(host: str, count: int = 4) -> dict:
    try:
        flag = "-n" if os.name == "nt" else "-c"
        r = subprocess.run(["ping", flag, str(count), host],
                           capture_output=True, text=True, timeout=30)
        return {"ok": r.returncode == 0, "stdout": (r.stdout or "")[:4000]}
    except Exception as e:
        return {"ok": False, "error": str(e)}


def web_search(query: str, max_results: int = 8) -> dict:
    """DuckDuckGo HTML search - no API key needed."""
    try:
        r = requests.post(
            "https://html.duckduckgo.com/html/",
            data={"q": query},
            headers={"User-Agent": "Mozilla/5.0 Ember"},
            timeout=15,
        )
        import re
        results = []
        for m in re.finditer(
            r'<a[^>]+class="result__a"[^>]+href="([^"]+)"[^>]*>(.*?)</a>'
            r'.*?<a[^>]+class="result__snippet"[^>]*>(.*?)</a>',
            r.text, re.DOTALL,
        ):
            url = urllib.parse.unquote(m.group(1))
            title = re.sub(r"<[^>]+>", "", m.group(2)).strip()
            snippet = re.sub(r"<[^>]+>", "", m.group(3)).strip()
            results.append({"title": title[:200], "url": url[:300], "snippet": snippet[:300]})
            if len(results) >= max_results:
                break
        return {"ok": True, "query": query, "result_count": len(results), "results": results}
    except Exception as e:
        return {"ok": False, "error": str(e)}


def wikipedia_summary(topic: str, sentences: int = 3) -> dict:
    """Free Wikipedia REST API."""
    try:
        title = urllib.parse.quote(topic.replace(" ", "_"))
        r = requests.get(f"https://en.wikipedia.org/api/rest_v1/page/summary/{title}", timeout=10)
        if r.status_code != 200:
            return {"ok": False, "error": f"HTTP {r.status_code}"}
        d = r.json()
        return {
            "ok": True, "title": d.get("title"),
            "extract": d.get("extract"),
            "url": d.get("content_urls", {}).get("desktop", {}).get("page"),
        }
    except Exception as e:
        return {"ok": False, "error": str(e)}


def weather_lookup(latitude: float, longitude: float) -> dict:
    """Free open-meteo - no API key. Pass lat/lon (decimal degrees)."""
    try:
        r = requests.get(
            "https://api.open-meteo.com/v1/forecast",
            params={
                "latitude": latitude, "longitude": longitude,
                "current": "temperature_2m,weather_code,wind_speed_10m,relative_humidity_2m",
                "timezone": "auto",
            },
            timeout=10,
        )
        d = r.json()
        return {"ok": True, **(d if isinstance(d, dict) else {"raw": d})}
    except Exception as e:
        return {"ok": False, "error": str(e)}


def translate_text(text: str, target_lang: str = "en", source_lang: str = "auto") -> dict:
    """Free Google Translate gtx endpoint - unofficial but stable for casual use."""
    try:
        r = requests.get(
            "https://translate.googleapis.com/translate_a/single",
            params={"client": "gtx", "sl": source_lang, "tl": target_lang, "dt": "t", "q": text},
            headers={"User-Agent": "Mozilla/5.0"},
            timeout=10,
        )
        data = r.json()
        translated = "".join(seg[0] for seg in data[0] if seg and seg[0])
        return {"ok": True, "source_lang": source_lang, "target_lang": target_lang,
                "input": text[:200], "output": translated}
    except Exception as e:
        return {"ok": False, "error": str(e)}


# ---------------------------------------------------------------------------
# Email
# ---------------------------------------------------------------------------
def send_email(to: str, subject: str, body: str, smtp_host: str | None = None,
               smtp_port: int = 587, smtp_user: str | None = None,
               smtp_password: str | None = None, html: bool = False) -> dict:
    """Send via SMTP. Credentials come from settings.json (email_smtp_* keys) if omitted.
    For Gmail: smtp.gmail.com:587, use an App Password (not your normal one)."""
    import smtplib

    try:
        from ui import load_settings
        st = load_settings()
        smtp_host = smtp_host or st.get("email_smtp_host", "")
        smtp_user = smtp_user or st.get("email_smtp_user", "")
        smtp_password = smtp_password or st.get("email_smtp_password", "")
        smtp_port = smtp_port or int(st.get("email_smtp_port", 587))
    except Exception:
        pass

    if not (smtp_host and smtp_user and smtp_password):
        return {"ok": False, "error": "SMTP not configured - set email_smtp_host/user/password in Settings"}

    try:
        msg = MIMEMultipart("alternative")
        msg["Subject"] = subject
        msg["From"] = smtp_user
        msg["To"] = to
        msg.attach(MIMEText(body, "html" if html else "plain"))
        with smtplib.SMTP(smtp_host, smtp_port, timeout=30) as s:
            s.starttls()
            s.login(smtp_user, smtp_password)
            s.sendmail(smtp_user, [to], msg.as_string())
        return {"ok": True, "to": to, "subject": subject}
    except Exception as e:
        return {"ok": False, "error": str(e)}


# ---------------------------------------------------------------------------
# Documents
# ---------------------------------------------------------------------------
def pdf_extract_text(path: str, max_pages: int = 50) -> dict:
    """Extract text from a PDF file (no OCR, just embedded text)."""
    try:
        from PyPDF2 import PdfReader
        reader = PdfReader(str(Path(path).expanduser()))
        total = len(reader.pages)
        out = []
        for i, pg in enumerate(reader.pages[:max_pages]):
            try:
                out.append(f"--- page {i + 1} ---\n" + (pg.extract_text() or "").strip())
            except Exception:
                continue
        return {"ok": True, "page_count": total, "extracted_pages": min(total, max_pages),
                "text": ("\n\n".join(out))[:50000]}
    except Exception as e:
        return {"ok": False, "error": str(e)}


def excel_read(path: str, sheet: str | None = None, max_rows: int = 200) -> dict:
    """Read .xlsx into a list of rows."""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(str(Path(path).expanduser()), read_only=True, data_only=True)
        ws = wb[sheet] if sheet else wb.active
        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= max_rows:
                break
            rows.append(list(row))
        return {"ok": True, "sheet": ws.title, "row_count": len(rows), "rows": rows}
    except Exception as e:
        return {"ok": False, "error": str(e)}


def excel_write(path: str, rows: list, sheet: str = "Sheet1") -> dict:
    """Write a list of rows (each a list of cells) to an .xlsx file."""
    try:
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = sheet
        for r in rows or []:
            ws.append(list(r) if isinstance(r, (list, tuple)) else [r])
        p = Path(path).expanduser()
        p.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(p))
        return {"ok": True, "path": str(p), "row_count": len(rows or [])}
    except Exception as e:
        return {"ok": False, "error": str(e)}


# ---------------------------------------------------------------------------
# Math / utility
# ---------------------------------------------------------------------------
_CALC_BINOPS = {
    ast.Add: operator.add, ast.Sub: operator.sub, ast.Mult: operator.mul,
    ast.Div: operator.truediv, ast.FloorDiv: operator.floordiv, ast.Mod: operator.mod,
    ast.Pow: operator.pow,
}
_CALC_UNARY = {ast.UAdd: operator.pos, ast.USub: operator.neg}
_CALC_FUNCS = {k: getattr(math, k) for k in dir(math)
               if not k.startswith("_") and callable(getattr(math, k))}
_CALC_FUNCS.update({"abs": abs, "round": round, "min": min, "max": max, "sum": sum, "pow": pow})
_CALC_CONSTS = {"pi": math.pi, "e": math.e, "tau": math.tau, "inf": math.inf, "nan": math.nan}


def _calc_eval(node):
    """Evaluate a parsed math AST against a strict whitelist — no eval(), so the old
    `().__class__` / `__import__` introspection escapes are structurally impossible."""
    if isinstance(node, ast.Expression):
        return _calc_eval(node.body)
    if isinstance(node, ast.Constant):
        if isinstance(node.value, bool) or not isinstance(node.value, (int, float)):
            raise ValueError("only numeric constants allowed")
        return node.value
    if isinstance(node, ast.UnaryOp) and type(node.op) in _CALC_UNARY:
        return _CALC_UNARY[type(node.op)](_calc_eval(node.operand))
    if isinstance(node, ast.BinOp) and type(node.op) in _CALC_BINOPS:
        left, right = _calc_eval(node.left), _calc_eval(node.right)
        if isinstance(node.op, ast.Pow) and abs(left) > 1 and abs(right) > 300:
            raise ValueError("exponent too large")  # block 9**9**9-style CPU/memory bombs
        return _CALC_BINOPS[type(node.op)](left, right)
    if isinstance(node, ast.Name):
        if node.id in _CALC_CONSTS:
            return _CALC_CONSTS[node.id]
        raise ValueError(f"unknown name: {node.id}")
    if isinstance(node, ast.Call):
        if not isinstance(node.func, ast.Name) or node.func.id not in _CALC_FUNCS or node.keywords:
            raise ValueError("only plain math functions allowed")
        if node.func.id == "factorial" and node.args and isinstance(node.args[0], ast.Constant) \
                and isinstance(node.args[0].value, (int, float)) and node.args[0].value > 10000:
            raise ValueError("factorial argument too large")
        return _CALC_FUNCS[node.func.id](*[_calc_eval(a) for a in node.args])
    raise ValueError("unsupported expression")


def calculator(expression: str) -> dict:
    """Safely evaluate a math expression. Supports +-*/%//**, parens, math.* functions.
    Uses an AST whitelist (not eval), so attribute access, subscripting, and builtins
    are rejected — no sandbox-escape via introspection."""
    try:
        tree = ast.parse(expression, mode="eval")
        val = _calc_eval(tree)
        return {"ok": True, "result": val, "expression": expression}
    except Exception as e:
        return {"ok": False, "error": str(e)}


def generate_password(length: int = 16, include_symbols: bool = True) -> dict:
    alphabet = string.ascii_letters + string.digits
    if include_symbols:
        alphabet += "!@#$%^&*()-_=+[]{};:,.<>?"
    pw = "".join(secrets.choice(alphabet) for _ in range(max(4, min(128, length))))
    return {"ok": True, "password": pw, "length": len(pw)}


def generate_uuid() -> dict:
    return {"ok": True, "uuid": str(uuid.uuid4())}


def hash_text(text: str, algorithm: str = "sha256") -> dict:
    try:
        h = hashlib.new(algorithm)
        h.update(text.encode("utf-8"))
        return {"ok": True, "algorithm": algorithm, "hex": h.hexdigest()}
    except Exception as e:
        return {"ok": False, "error": str(e)}


def hash_file(path: str, algorithm: str = "sha256") -> dict:
    try:
        h = hashlib.new(algorithm)
        with Path(path).expanduser().open("rb") as f:
            for chunk in iter(lambda: f.read(1 << 20), b""):
                h.update(chunk)
        return {"ok": True, "algorithm": algorithm, "hex": h.hexdigest(), "path": str(path)}
    except Exception as e:
        return {"ok": False, "error": str(e)}


def base64_encode(text: str) -> dict:
    return {"ok": True, "encoded": base64.b64encode(text.encode("utf-8")).decode("ascii")}


def base64_decode(text: str) -> dict:
    try:
        return {"ok": True, "decoded": base64.b64decode(text).decode("utf-8", errors="replace")}
    except Exception as e:
        return {"ok": False, "error": str(e)}


def url_encode(text: str) -> dict:
    return {"ok": True, "encoded": urllib.parse.quote(text, safe="")}


def url_decode(text: str) -> dict:
    return {"ok": True, "decoded": urllib.parse.unquote(text)}


def now(timezone_name: str = "local") -> dict:
    if timezone_name == "local":
        dt = datetime.now()
    elif timezone_name in ("utc", "UTC"):
        dt = datetime.now(timezone.utc)
    else:
        try:
            from zoneinfo import ZoneInfo
            dt = datetime.now(ZoneInfo(timezone_name))
        except Exception as e:
            return {"ok": False, "error": str(e)}
    return {
        "ok": True,
        "iso": dt.isoformat(timespec="seconds"),
        "epoch": int(dt.timestamp()) if dt.tzinfo else int(datetime.now().timestamp()),
        "weekday": dt.strftime("%A"),
        "date": dt.strftime("%Y-%m-%d"),
        "time": dt.strftime("%H:%M:%S"),
        "timezone": timezone_name,
    }


# ---------------------------------------------------------------------------
# System / hardware
# ---------------------------------------------------------------------------
def power_action(action: str, force: bool = False) -> dict:
    """Lock / sleep / restart / shutdown / hibernate / logoff the PC. Windows only.
    USE WITH CARE - shutdown/restart will close unsaved work."""
    if os.name != "nt":
        return {"ok": False, "error": "Windows only"}
    cmds = {
        "lock": "rundll32.exe user32.dll,LockWorkStation",
        "sleep": "rundll32.exe powrprof.dll,SetSuspendState 0,1,0",
        "hibernate": "shutdown /h",
        "restart": f"shutdown /r /t 0{' /f' if force else ''}",
        "shutdown": f"shutdown /s /t 0{' /f' if force else ''}",
        "logoff": "shutdown /l",
    }
    if action not in cmds:
        return {"ok": False, "error": f"unknown action; pick from {list(cmds)}"}
    try:
        subprocess.Popen(cmds[action], shell=True)
        return {"ok": True, "action": action}
    except Exception as e:
        return {"ok": False, "error": str(e)}


def get_battery() -> dict:
    try:
        import psutil
        b = psutil.sensors_battery()
        if not b:
            return {"ok": True, "has_battery": False}
        return {
            "ok": True, "has_battery": True,
            "percent": round(b.percent, 1),
            "plugged_in": b.power_plugged,
            "minutes_left": (b.secsleft // 60) if b.secsleft not in (psutil.POWER_TIME_UNLIMITED, psutil.POWER_TIME_UNKNOWN) else None,
        }
    except Exception as e:
        return {"ok": False, "error": str(e)}


def _osascript(script: str) -> str:
    """Run an AppleScript one-liner and return stripped stdout (mac only)."""
    return subprocess.run(["osascript", "-e", script], capture_output=True,
                          text=True, timeout=10).stdout.strip()


def get_volume() -> dict:
    """Master volume level 0-100 (+ muted state)."""
    if sys.platform == "darwin":
        try:
            out = _osascript("output volume of (get volume settings)")
            muted = _osascript("output muted of (get volume settings)")
            return {"ok": True, "percent": float(out), "muted": muted.lower() == "true"}
        except Exception as e:
            return {"ok": False, "error": str(e)}
    try:
        from ctypes import POINTER, cast
        from comtypes import CLSCTX_ALL
        from pycaw.pycaw import AudioUtilities, IAudioEndpointVolume
        device = AudioUtilities.GetSpeakers()
        interface = device.Activate(IAudioEndpointVolume._iid_, CLSCTX_ALL, None)
        vol = cast(interface, POINTER(IAudioEndpointVolume))
        scalar = vol.GetMasterVolumeLevelScalar()
        muted = vol.GetMute()
        return {"ok": True, "percent": round(scalar * 100, 1), "muted": bool(muted)}
    except Exception as e:
        return {"ok": False, "error": str(e)}


def set_volume(percent: float) -> dict:
    """Set master volume 0-100."""
    pct = max(0.0, min(100.0, float(percent)))
    if sys.platform == "darwin":
        try:
            _osascript(f"set volume output volume {int(round(pct))}")
            return {"ok": True, "percent": pct}
        except Exception as e:
            return {"ok": False, "error": str(e)}
    try:
        from ctypes import POINTER, cast
        from comtypes import CLSCTX_ALL
        from pycaw.pycaw import AudioUtilities, IAudioEndpointVolume
        device = AudioUtilities.GetSpeakers()
        interface = device.Activate(IAudioEndpointVolume._iid_, CLSCTX_ALL, None)
        vol = cast(interface, POINTER(IAudioEndpointVolume))
        scalar = pct / 100.0
        vol.SetMasterVolumeLevelScalar(scalar, None)
        return {"ok": True, "percent": round(scalar * 100, 1)}
    except Exception as e:
        return {"ok": False, "error": str(e)}


def toggle_mute() -> dict:
    if sys.platform == "darwin":
        try:
            cur = _osascript("output muted of (get volume settings)").lower() == "true"
            _osascript(f"set volume {'without' if cur else 'with'} output muted")
            return {"ok": True, "muted": not cur}
        except Exception as e:
            return {"ok": False, "error": str(e)}
    try:
        from ctypes import POINTER, cast
        from comtypes import CLSCTX_ALL
        from pycaw.pycaw import AudioUtilities, IAudioEndpointVolume
        device = AudioUtilities.GetSpeakers()
        interface = device.Activate(IAudioEndpointVolume._iid_, CLSCTX_ALL, None)
        vol = cast(interface, POINTER(IAudioEndpointVolume))
        cur = vol.GetMute()
        vol.SetMute(0 if cur else 1, None)
        return {"ok": True, "muted": not cur}
    except Exception as e:
        return {"ok": False, "error": str(e)}


def env_get(name: str) -> dict:
    return {"ok": True, "name": name, "value": os.environ.get(name)}


def env_list(prefix: str = "") -> dict:
    items = {k: v for k, v in os.environ.items() if (not prefix or k.startswith(prefix))}
    return {"ok": True, "count": len(items), "vars": items}


# ---------------------------------------------------------------------------
# Image processing
# ---------------------------------------------------------------------------
def image_resize(path: str, max_dimension: int = 1024, destination: str | None = None,
                 quality: int = 85) -> dict:
    try:
        from PIL import Image
        src = Path(path).expanduser()
        img = Image.open(src)
        w, h = img.size
        if max(w, h) > max_dimension:
            ratio = max_dimension / max(w, h)
            img = img.resize((int(w * ratio), int(h * ratio)), Image.LANCZOS)
        dst = Path(destination).expanduser() if destination else src.with_name(f"{src.stem}_resized{src.suffix}")
        save_kwargs = {"quality": quality, "optimize": True} if src.suffix.lower() in (".jpg", ".jpeg") else {}
        img.save(str(dst), **save_kwargs)
        return {"ok": True, "path": str(dst), "new_size": list(img.size), "original_size": [w, h]}
    except Exception as e:
        return {"ok": False, "error": str(e)}


def image_crop(path: str, left: int, top: int, right: int, bottom: int,
               destination: str | None = None) -> dict:
    try:
        from PIL import Image
        src = Path(path).expanduser()
        img = Image.open(src).crop((left, top, right, bottom))
        dst = Path(destination).expanduser() if destination else src.with_name(f"{src.stem}_crop{src.suffix}")
        img.save(str(dst))
        return {"ok": True, "path": str(dst), "size": list(img.size)}
    except Exception as e:
        return {"ok": False, "error": str(e)}


def image_convert(path: str, output_format: str, destination: str | None = None) -> dict:
    try:
        from PIL import Image
        src = Path(path).expanduser()
        img = Image.open(src)
        out_format = output_format.lower().lstrip(".")
        if out_format in ("jpg", "jpeg") and img.mode in ("RGBA", "P"):
            img = img.convert("RGB")
        dst = Path(destination).expanduser() if destination else src.with_suffix(f".{out_format}")
        img.save(str(dst), format=out_format.upper().replace("JPG", "JPEG"))
        return {"ok": True, "path": str(dst), "format": out_format}
    except Exception as e:
        return {"ok": False, "error": str(e)}


def color_at(x: int, y: int) -> dict:
    """Read the RGB color of the pixel at logical screen coordinates (x, y) — the same
    coordinate space click() uses. Converts to physical pixels so the sample is correct
    on Retina displays (previously it read an offset pixel)."""
    try:
        import mss
        from PIL import Image
        try:
            import screen_vision
            sx, sy, mon = screen_vision._scale()
            px = int(mon["left"] + int(x) * sx)
            py = int(mon["top"] + int(y) * sy)
        except Exception:
            px, py = int(x), int(y)
        with mss.mss() as sct:
            raw = sct.grab({"left": px, "top": py, "width": 1, "height": 1})
            img = Image.frombytes("RGB", raw.size, raw.bgra, "raw", "BGRX")
            r, g, b = img.getpixel((0, 0))
        return {"ok": True, "x": x, "y": y, "rgb": [r, g, b],
                "hex": f"#{r:02x}{g:02x}{b:02x}"}
    except Exception as e:
        return {"ok": False, "error": str(e)}


def qr_generate(text: str, destination: str | None = None, size: int = 6) -> dict:
    """Generate a QR-code PNG."""
    try:
        import qrcode
        img = qrcode.make(text, box_size=size, border=2)
        dst = Path(destination).expanduser() if destination else Path.home() / "Downloads" / f"qr_{int(time.time())}.png"
        dst.parent.mkdir(parents=True, exist_ok=True)
        img.save(str(dst))
        return {"ok": True, "path": str(dst), "encoded": text}
    except Exception as e:
        return {"ok": False, "error": str(e)}


# ---------------------------------------------------------------------------
# Clipboard
# ---------------------------------------------------------------------------
_clipboard_history: list[str] = []


def clipboard_get() -> dict:
    try:
        import pyperclip
        v = pyperclip.paste()
        return {"ok": True, "text": v[:5000], "length": len(v)}
    except Exception as e:
        return {"ok": False, "error": str(e)}


def clipboard_set(text: str) -> dict:
    try:
        import pyperclip
        pyperclip.copy(text)
        return {"ok": True, "length": len(text)}
    except Exception as e:
        return {"ok": False, "error": str(e)}


def clipboard_history_get(max_items: int = 20) -> dict:
    """Returns the in-process clipboard ring. Populated by snapshot calls."""
    return {"ok": True, "count": len(_clipboard_history),
            "items": _clipboard_history[-max_items:]}


def clipboard_history_snapshot() -> dict:
    """Capture the current clipboard into the in-process history ring."""
    try:
        import pyperclip
        v = pyperclip.paste()
        if v and (not _clipboard_history or _clipboard_history[-1] != v):
            _clipboard_history.append(v[:5000])
            if len(_clipboard_history) > 100:
                del _clipboard_history[0]
        return {"ok": True, "snapshot_count": len(_clipboard_history)}
    except Exception as e:
        return {"ok": False, "error": str(e)}


# ---------------------------------------------------------------------------
# Git
# ---------------------------------------------------------------------------
def git_status(repo: str = ".") -> dict:
    try:
        r = subprocess.run(["git", "-C", str(Path(repo).expanduser()), "status", "--short", "--branch"],
                           capture_output=True, text=True, timeout=10)
        return {"ok": r.returncode == 0, "output": (r.stdout or r.stderr)[:4000]}
    except Exception as e:
        return {"ok": False, "error": str(e)}


def git_log(repo: str = ".", count: int = 20) -> dict:
    try:
        r = subprocess.run(
            ["git", "-C", str(Path(repo).expanduser()), "log",
             f"-{count}", "--pretty=format:%h %an %ar  %s"],
            capture_output=True, text=True, timeout=10)
        return {"ok": r.returncode == 0, "log": (r.stdout or "")[:6000]}
    except Exception as e:
        return {"ok": False, "error": str(e)}


def git_diff(repo: str = ".", staged: bool = False) -> dict:
    try:
        args = ["git", "-C", str(Path(repo).expanduser()), "diff"]
        if staged:
            args.append("--staged")
        r = subprocess.run(args, capture_output=True, text=True, timeout=15)
        return {"ok": r.returncode == 0, "diff": (r.stdout or "")[:10000]}
    except Exception as e:
        return {"ok": False, "error": str(e)}


# ---------------------------------------------------------------------------
# Calendar (.ics file output - works with any calendar app)
# ---------------------------------------------------------------------------
def create_calendar_event(title: str, start: str, end: str | None = None,
                           description: str = "", location: str = "",
                           destination: str | None = None) -> dict:
    """Build an .ics calendar event file. start/end ISO 8601 e.g. 2026-06-01T15:00:00."""
    try:
        s = datetime.fromisoformat(start.replace("Z", "+00:00"))
        e = datetime.fromisoformat(end.replace("Z", "+00:00")) if end else s + timedelta(hours=1)
        ics = (
            "BEGIN:VCALENDAR\r\nVERSION:2.0\r\nPRODID:-//Ember//EN\r\n"
            "BEGIN:VEVENT\r\n"
            f"UID:{uuid.uuid4()}@ember\r\n"
            f"DTSTAMP:{datetime.utcnow().strftime('%Y%m%dT%H%M%SZ')}\r\n"
            f"DTSTART:{s.strftime('%Y%m%dT%H%M%S')}\r\n"
            f"DTEND:{e.strftime('%Y%m%dT%H%M%S')}\r\n"
            f"SUMMARY:{title}\r\n"
            f"DESCRIPTION:{description}\r\n"
            f"LOCATION:{location}\r\n"
            "END:VEVENT\r\nEND:VCALENDAR\r\n"
        )
        dst = Path(destination).expanduser() if destination else Path.home() / "Downloads" / f"event_{int(time.time())}.ics"
        dst.parent.mkdir(parents=True, exist_ok=True)
        dst.write_text(ics, encoding="utf-8")
        return {"ok": True, "path": str(dst), "title": title, "starts": s.isoformat()}
    except Exception as e:
        return {"ok": False, "error": str(e)}


# ---------------------------------------------------------------------------
# JSON / CSV
# ---------------------------------------------------------------------------
def json_query(path: str, dotted_path: str = "") -> dict:
    """Read a JSON file and pull out a value via dotted path (e.g. 'users.0.name')."""
    try:
        data = json.loads(Path(path).expanduser().read_text(encoding="utf-8"))
        if not dotted_path:
            return {"ok": True, "value": data}
        cur = data
        for piece in dotted_path.split("."):
            if isinstance(cur, list) and piece.lstrip("-").isdigit():
                cur = cur[int(piece)]
            elif isinstance(cur, dict):
                cur = cur.get(piece)
            else:
                cur = None
                break
        return {"ok": True, "path": dotted_path, "value": cur}
    except Exception as e:
        return {"ok": False, "error": str(e)}


def csv_read(path: str, max_rows: int = 200, delimiter: str = ",") -> dict:
    try:
        import csv as _csv
        with Path(path).expanduser().open("r", encoding="utf-8", newline="") as f:
            reader = _csv.reader(f, delimiter=delimiter)
            rows = []
            for i, r in enumerate(reader):
                if i >= max_rows:
                    break
                rows.append(r)
        return {"ok": True, "row_count": len(rows), "rows": rows}
    except Exception as e:
        return {"ok": False, "error": str(e)}
